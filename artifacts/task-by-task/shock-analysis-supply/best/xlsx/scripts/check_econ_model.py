#!/usr/bin/env python3
"""
check_econ_model.py -- magnitude/units sanity checker for multi-source
economic models (production-function / potential-GDP / Solow-TFP builds).

WHY THIS EXISTS
A log-linear TFP residual (lnZ = lnY - alpha*lnK) silently ABSORBS any constant
scale offset between capital (K) and output (Y). So a model can recalculate with
ZERO formula errors and clean LnK/LnY/LnZ columns while every headline VALUE is
off by a round factor (x1000) because one series was left in billions and the
other in millions. The recalc "0 errors" check cannot catch this. This script
catches it by testing economic INVARIANTS on the calculated values.

Run this on the recalculated workbook AFTER building the model and BEFORE
declaring the task done:

    python scripts/check_econ_model.py <file.xlsx>

It reads calculated values (data_only), auto-locates the capital ("K") and output
("Real GDP"/"GDP"/"Y") columns by their header labels, and reports:

  * capital-output ratio  K/Y  -- must be a small number (roughly 1-30 for any
    real economy). A ratio in the thousands means K and Y are in mismatched
    scales -> scale the smaller series (usually multiply the GDP link by 1000
    to turn billions into millions) so both share ONE reporting unit.
  * depreciation rate      -- an annual capital depreciation rate is a few
    percent (~1-10%). A value below ~0.8% or above ~30% signals the CFC-flow and
    capital-stock series are in mismatched units/currencies.

Exit code is non-zero when a likely units error is detected, so it is easy to
gate on. Nothing is written; the workbook is only read.
"""
import sys
import statistics

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required (pip install openpyxl).")
    sys.exit(2)

# Header label synonyms (case-insensitive, stripped).
CAPITAL_LABELS = {"k", "capital", "capital stock", "capital_stock"}
OUTPUT_LABELS = {"real gdp", "gdp", "y", "output", "real output", "realgdp"}

# Plausible bands (general economic invariants, NOT task thresholds).
KY_LOW, KY_HIGH = 0.3, 60.0            # capital-output ratio for any economy
DEP_LOW, DEP_HIGH = 0.008, 0.30        # annual depreciation rate (0.8%-30%)


def _norm(v):
    return str(v).strip().lower() if v is not None else ""


def find_header_columns(ws, max_header_row=12):
    """Return (header_row, {capital_col, output_col}) if a row contains both a
    capital label and an output label; else None."""
    for r in range(1, max_header_row + 1):
        cap_col = out_col = None
        for c in range(1, ws.max_column + 1):
            lab = _norm(ws.cell(row=r, column=c).value)
            if not lab:
                continue
            if lab in CAPITAL_LABELS and cap_col is None:
                cap_col = c
            elif lab in OUTPUT_LABELS and out_col is None:
                out_col = c
        if cap_col and out_col:
            return r, cap_col, out_col
    return None


def collect_ratios(ws, header_row, cap_col, out_col, max_rows=200):
    ratios = []
    for r in range(header_row + 1, header_row + 1 + max_rows):
        K = ws.cell(row=r, column=cap_col).value
        Y = ws.cell(row=r, column=out_col).value
        if isinstance(K, (int, float)) and isinstance(Y, (int, float)) and Y:
            if K > 0 and Y > 0:
                ratios.append(K / Y)
    return ratios


def find_depreciation(wb):
    """Find a cell labelled like a depreciation rate and return its neighbour's
    numeric value (value in a cell to the right of or below the label)."""
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 60) + 1):
            for c in range(1, min(ws.max_column, 20) + 1):
                lab = _norm(ws.cell(row=r, column=c).value)
                if "depreciation" in lab:
                    for rr, cc in ((r, c + 1), (r + 1, c), (r, c + 2)):
                        v = ws.cell(row=rr, column=cc).value
                        if isinstance(v, (int, float)):
                            return ws.title, rr, cc, v
    return None


def main():
    if len(sys.argv) < 2:
        print("usage: python check_econ_model.py <file.xlsx>")
        sys.exit(2)
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)

    problems = []
    checked_ky = False

    for ws in wb.worksheets:
        hit = find_header_columns(ws)
        if not hit:
            continue
        header_row, cap_col, out_col = hit
        ratios = collect_ratios(ws, header_row, cap_col, out_col)
        if len(ratios) < 3:
            continue
        checked_ky = True
        med = statistics.median(ratios)
        col = openpyxl.utils.get_column_letter
        print(f"[{ws.title}] capital-output ratio K/Y: median={med:.3g} "
              f"(n={len(ratios)}, K=col {col(cap_col)}, Y=col {col(out_col)})")
        if med > KY_HIGH:
            factor = 1000 if med > 200 else round(med / 5)
            problems.append(
                f"UNITS MISMATCH in '{ws.title}': K/Y median={med:.4g} is far "
                f"above the plausible 1-30 range. Capital and output are in "
                f"DIFFERENT scales. Your output (GDP) series is ~{factor}x too "
                f"small relative to capital -- almost certainly GDP was left in "
                f"BILLIONS while capital is in MILLIONS. Multiply the GDP link "
                f"by 1000 (e.g. ='WEO_Data'!C10*1000) so both use ONE unit, then "
                f"recalc. In a log-linear TFP model this also rescales every "
                f"downstream Ystar/potential-output value by the same factor."
            )
        elif med < KY_LOW:
            problems.append(
                f"UNITS MISMATCH in '{ws.title}': K/Y median={med:.4g} is far "
                f"below the plausible 1-30 range. Capital is ~{round(1/med)}x too "
                f"small relative to output -- rescale capital (or GDP) so both "
                f"share ONE reporting unit, then recalc."
            )
        else:
            print(f"  -> OK: within the plausible 1-30 capital-output band.")

    dep = find_depreciation(wb)
    if dep:
        sheet, r, c, v = dep
        col = openpyxl.utils.get_column_letter(c)
        print(f"[{sheet}] depreciation rate {col}{r}: {v:.4g} ({v*100:.2f}%)")
        if not (DEP_LOW <= v <= DEP_HIGH):
            problems.append(
                f"IMPLAUSIBLE DEPRECIATION RATE in '{sheet}' {col}{r}={v:.4g} "
                f"({v*100:.2f}%). A real annual capital depreciation rate is a "
                f"few percent (~1-10%). A value this far outside 0.8%-30% means "
                f"the CFC (consumption-of-fixed-capital) flow and the capital "
                f"stock it is divided by are in MISMATCHED units/currencies. "
                f"Reconcile them to one unit (rate = CFC / capital-stock, both in "
                f"the SAME currency and scale) rather than accepting the number."
            )
        else:
            print(f"  -> OK: within the plausible ~1-10% depreciation band.")

    print()
    if not checked_ky and not dep:
        print("NOTE: could not locate a capital ('K') column + output "
              "('Real GDP'/'GDP') column, nor a depreciation label. Build the "
              "model (with those labelled columns populated & recalculated) "
              "before running this check.")
        sys.exit(0)

    if problems:
        print("=" * 70)
        print(f"FAIL: {len(problems)} likely units/magnitude error(s) detected:")
        for p in problems:
            print("  * " + p)
        print("=" * 70)
        print("Do NOT finish the task while this reports FAIL -- fix the scale "
              "factor(s), recalc, and re-run this check until it reports OK.")
        sys.exit(1)

    print("OK: magnitudes are self-consistent (capital-output ratio and "
          "depreciation rate are in plausible economic ranges).")
    sys.exit(0)


if __name__ == "__main__":
    main()
