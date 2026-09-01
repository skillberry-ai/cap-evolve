#!/usr/bin/env python3
"""
Cached-Value Verifier for Excel files.

openpyxl saves formulas as text but does NOT compute them, so a workbook that
was written (or re-saved) by openpyxl contains formulas with NO cached result.
Anything that reads the file with cached values only -- openpyxl
`data_only=True`, pandas `read_excel`, most graders/verifiers -- then sees
`None`/blank for those cells, not the computed number.

`recalc.py` (LibreOffice) is what writes the cached values back in. This script
checks whether the file you are about to deliver actually HAS cached values for
its formulas. Run it as the FINAL step, AFTER recalc.py, on the real delivered
file (not a throwaway copy). If it reports `stale`, re-run recalc.py on THIS
file before finishing.

Usage:
    python verify_values.py <excel_file>

Exit code 0 => ok (every formula has a cached value).
Exit code 1 => stale (some formulas have no cached value; re-run recalc.py) or error.
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def verify(filename):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}, 1

    wb_f = load_workbook(filename, data_only=False)
    wb_v = load_workbook(filename, data_only=True)

    total_formulas = 0
    missing = []  # formula cells with no cached value
    for name in wb_f.sheetnames:
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        for row in ws_f.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    total_formulas += 1
                    cached = ws_v[cell.coordinate].value
                    if cached is None:
                        missing.append(f"{name}!{cell.coordinate}")
    wb_f.close()
    wb_v.close()

    if missing:
        result = {
            "status": "stale",
            "total_formulas": total_formulas,
            "formulas_missing_cached_value": len(missing),
            "locations": missing[:20],
            "hint": "Formulas have no cached values (file was saved by openpyxl "
                    "without recalculating). Re-run: python recalc.py " + str(filename),
        }
        return result, 1

    return {"status": "ok", "total_formulas": total_formulas,
            "formulas_missing_cached_value": 0}, 0


def main():
    if len(sys.argv) < 2:
        print("Usage: python verify_values.py <excel_file>")
        sys.exit(1)
    result, code = verify(sys.argv[1])
    print(json.dumps(result, indent=2))
    sys.exit(code)


if __name__ == "__main__":
    main()
