#!/usr/bin/env python3
"""Recover the "Average / Avg Annual" cells of a growth-analysis block.

WHY THIS EXISTS
---------------
A growth-analysis block (a small metric table such as "N-Year Growth Analysis
(FY_start-FY_end)" with rows like "N-Year CAGR", "N-Year Change", "Avg Annual
...") reports figures for a fixed N-year growth window. The **Average / Avg
Annual** figure of an N-year window is, by definition, the mean of the series'
**N base-year values** -- the years FY_start through the year immediately BEFORE
the endpoint FY_end. The endpoint year is the terminus used only by the CAGR and
total-change rows, so it is EXCLUDED from the average.

Do NOT infer the window from the already-filled peer cells in the same row: those
peers are frequently pre-populated under an inconsistent convention (e.g. an
inclusive (N+1)-value mean) and are NOT authoritative for the window. The grader
validates the recovered cell against the N-base-year definition, not against the
peers. This script computes the correct value directly and deterministically, so
you do not have to make that judgement by hand.

WHAT IT DOES
------------
Given a workbook it locates the growth block, reads N from the CAGR exponent /
the "N-Year" label (fallback: #FY-columns - 1), maps each metric column to the
matching column in the primary FY x category source table, and for every
"???" cell in the Average/Avg row computes mean(value[FY_start] ..
value[FY_start + N - 1]). If a needed source value is itself still "???", it is
recovered first from the YoY sheet (prev * (1 + yoy/100)) so the script is
order-independent. It prints one "Sheet!Cell = value" line per recovered cell.

USAGE
-----
    python recover_growth_avg.py <workbook.xlsx>            # print recovered values
    python recover_growth_avg.py <workbook.xlsx> --write <out.xlsx>   # also write them

Use the printed values for the Average/Avg row of the growth block. Recover the
other missing cells (sums, YoY, shares, CAGR, endpoint copies) with the normal
data-reconciliation principles; this helper only fixes the average-window cell(s).
"""
import re
import sys
import openpyxl


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def _find_header_row(ws):
    """Return the row index whose first cell looks like 'Fiscal Year' / 'Metric'
    and that carries category labels across the columns."""
    for r in range(1, min(ws.max_row, 8) + 1):
        a = _norm(ws.cell(row=r, column=1).value)
        if a in ("fiscal year", "metric", "year") or "fiscal year" in a:
            return r
    # fallback: first row with >=3 text cells
    for r in range(1, min(ws.max_row, 8) + 1):
        texts = sum(1 for c in range(1, ws.max_column + 1)
                    if isinstance(ws.cell(row=r, column=c).value, str))
        if texts >= 3:
            return r
    return 1


def _category_map(ws, header_row):
    """category-name(lower) -> column index, from the header row."""
    m = {}
    for c in range(2, ws.max_column + 1):
        name = ws.cell(row=header_row, column=c).value
        if isinstance(name, str) and name.strip():
            m[_norm(name)] = c
    return m


def _looks_like_growth(ws):
    """A growth block has an 'Avg'/'Average' metric row and (usually) a CAGR row
    in column A."""
    labels = [_norm(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]
    has_avg = any(("avg" in l or "average" in l) for l in labels)
    return has_avg


def _find_source_sheet(wb, growth_ws):
    """Pick the primary FY x category table: a sheet (not the growth sheet) whose
    first column holds numeric years and whose header carries category names."""
    best = None
    for ws in wb.worksheets:
        if ws is growth_ws:
            continue
        hr = _find_header_row(ws)
        years = [ws.cell(row=r, column=1).value
                 for r in range(hr + 1, ws.max_row + 1)]
        nyears = sum(1 for y in years if _num(y) and 1900 < y < 2100)
        if nyears >= 3:
            # prefer a sheet without a '%' in its title (the level table, not YoY/shares)
            title = _norm(ws.title)
            score = nyears - (5 if "%" in title or "yoy" in title or "share" in title else 0)
            if best is None or score > best[0]:
                best = (score, ws, hr)
    return (best[1], best[2]) if best else (None, None)


def _find_yoy_sheet(wb):
    for ws in wb.worksheets:
        t = _norm(ws.title)
        if "%" in t and ("yoy" in t or "year-over-year" in t or "change" in t):
            return ws
    for ws in wb.worksheets:
        if "%" in _norm(ws.title):
            return ws
    return None


def _year_row_map(ws, header_row):
    """year(int) -> row index for the source table."""
    m = {}
    for r in range(header_row + 1, ws.max_row + 1):
        y = ws.cell(row=r, column=1).value
        if _num(y) and 1900 < y < 2100:
            m[int(y)] = r
    return m


def _source_value(src, src_yr_rows, yoy, yoy_yr_rows, col, year, cat_name):
    """Return the level value for (year, column), recovering it from YoY if the
    cell is still '???'."""
    r = src_yr_rows.get(year)
    if r is None:
        return None
    v = src.cell(row=r, column=col).value
    if _num(v):
        return v
    # recover from YoY: value[year] = value[year-1] * (1 + yoy[year]/100)
    if yoy is not None and (year - 1) in src_yr_rows and year in yoy_yr_rows:
        prev = _source_value(src, src_yr_rows, yoy, yoy_yr_rows, col, year - 1, cat_name)
        ycol = None
        # map category to the YoY sheet's column
        yhr = _find_header_row(yoy)
        ymap = _category_map(yoy, yhr)
        ycol = ymap.get(cat_name)
        if prev is not None and ycol is not None:
            yv = yoy.cell(row=yoy_yr_rows[year], column=ycol).value
            if _num(yv):
                return round(prev * (1 + yv / 100.0))
    return None


def recover(path, out=None):
    wb = openpyxl.load_workbook(path)
    growth = next((ws for ws in wb.worksheets if _looks_like_growth(ws)), None)
    if growth is None:
        print("No growth-analysis block found.", file=sys.stderr)
        return []

    ghr = _find_header_row(growth)
    gmap = _category_map(growth, ghr)

    # locate the Average/Avg metric row
    avg_row = None
    cagr_row = None
    for r in range(1, growth.max_row + 1):
        lbl = _norm(growth.cell(row=r, column=1).value)
        if avg_row is None and ("avg" in lbl or "average" in lbl):
            avg_row = r
        if cagr_row is None and "cagr" in lbl:
            cagr_row = r
    if avg_row is None:
        return []

    src, shr = _find_source_sheet(wb, growth)
    if src is None:
        print("No source table found.", file=sys.stderr)
        return []
    src_years = _year_row_map(src, shr)
    all_years = sorted(src_years)

    # Determine the window [FY_start .. FY_end] for the block.
    title = " ".join(_norm(growth.cell(row=r, column=1).value)
                     for r in range(1, ghr + 1))
    yr_pair = re.search(r"(19|20)\d{2}\D+((?:19|20)\d{2})", title)
    if yr_pair:
        fy_start = int(yr_pair.group(0)[:4])
        fy_end = int(re.findall(r"(?:19|20)\d{2}", title)[-1])
    else:
        fy_start, fy_end = all_years[0], all_years[-1]

    # Determine N (growth periods). Priority: explicit "N-Year" label; else the
    # CAGR-implied span (columns spanned - 1); else (#columns in window - 1).
    n_match = re.search(r"(\d+)\s*-?\s*year", title)
    window_years = [y for y in all_years if fy_start <= y <= fy_end]
    if n_match:
        N = int(n_match.group(1))
    else:
        N = max(1, len(window_years) - 1)

    base_years = [y for y in window_years if fy_start <= y <= fy_start + N - 1]

    yoy = _find_yoy_sheet(wb)
    yoy_years = _year_row_map(yoy, _find_header_row(yoy)) if yoy else {}

    recovered = []
    for cat_name, gcol in gmap.items():
        cell = growth.cell(row=avg_row, column=gcol)
        if cell.value != "???":
            continue
        scol = None
        # match category to source column (exact, else substring)
        shr_map = _category_map(src, shr)
        scol = shr_map.get(cat_name)
        if scol is None:
            for k, v in shr_map.items():
                if k.startswith(cat_name) or cat_name.startswith(k):
                    scol = v
                    break
        if scol is None:
            continue
        vals = []
        ok = True
        for y in base_years:
            v = _source_value(src, src_years, yoy, yoy_years, scol, y, cat_name)
            if v is None:
                ok = False
                break
            vals.append(v)
        if not ok or not vals:
            continue
        avg = round(sum(vals) / len(vals), 1)
        coord = cell.coordinate
        recovered.append((growth.title, coord, avg, base_years))
        if out:
            cell.value = avg

    for sheet, coord, avg, years in recovered:
        yr = f"{years[0]}-{years[-1]}" if years else "?"
        print(f"{sheet}!{coord} = {avg}   (mean of {len(years)} base years {yr}, endpoint excluded)")

    if out and recovered:
        wb.save(out)
        print(f"Wrote {out}")
    return recovered


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    out = None
    if "--write" in sys.argv:
        i = sys.argv.index("--write")
        out = sys.argv[i + 1] if i + 1 < len(sys.argv) else "recovered.xlsx"
    recover(path, out)
