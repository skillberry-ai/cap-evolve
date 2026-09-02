#!/usr/bin/env python3
"""
Scaffold a macro-accounting economic **shock / impact analysis** workbook
(demand-side National-Accounts framework).

WHEN TO USE THIS
----------------
Use this for tasks that ask you to estimate an investment/spending shock to an
economy with the macro accounting (demand-side) framework and require an Excel
workbook containing these linked sheets:

  * ``WEO_Data``          - IMF-WEO style GDP time series with projected years
  * ``SUT Calc``          - supply-use table calculation + import-content share
  * ``SUPPLY (38-38)-2024`` / ``USE (38-38)-2024`` - the two Geostat SUT sheets
  * ``NA``                - National-Accounts impact table + scenario blocks

It builds the **exact sheet structure and formula wiring** these models need
(time index in column A/B, one variable per column, projected periods written as
chain formulas, cross-sheet links, scenario assumption blocks). It does **not**
invent any source data: the historical actuals (IMF real-GDP levels, deflator,
and the Geostat SUPPLY/USE cells) are left blank/placeholder for you to fill in
after you collect them from the authority websites. Because every calculated
cell is an Excel *formula* (never a Python-computed constant), the file stays
dynamic and recalculates correctly with ``recalc.py``.

WHY A SCRIPT
------------
Hand-building this workbook repeatedly puts columns in the wrong order, hardcodes
projected values instead of formulas, and drops the required sheet names -
deterministic mistakes. Running this scaffolder gets the structure right every
time; you then (1) paste the collected actuals into the blank cells and
(2) run ``recalc.py`` to evaluate the formulas.

USAGE
-----
    python scripts/build_shock_model.py --out test_demand.xlsx \
        --investment 6500 --fx 2.746 --multiplier 0.8 \
        --scenario2-multiplier 1.0 --scenario3-import-share 0.5

All numeric parameters come from the *task statement* (the USD investment size,
the Lari/USD FX rate, the small-open-economy demand multiplier, and the scenario
overrides). Defaults match a "6.5B USD, 8-year, bell-shaped allocation" demand
shock but every value is overridable, so the same scaffolder serves the whole
shock-analysis task family. Nothing here is a task answer key - only structure.

If ``--out`` already exists (e.g. a provided template that already contains the
``NA`` sheet with its year rows and assumption labels), that sheet is preserved
and filled in place; missing sheets are added.
"""
import argparse
from openpyxl import Workbook, load_workbook


def _bell(n):
    """Symmetric bell-shaped allocation weights over n periods, summing to ~1.
    For the canonical 8-period case returns the standard
    0.05, 0.10, 0.15, 0.20, 0.20, 0.15, 0.10, 0.05 profile."""
    if n == 8:
        return [0.05, 0.10, 0.15, 0.20, 0.20, 0.15, 0.10, 0.05]
    # generic symmetric triangular-ish bell for other horizons
    half = (n + 1) // 2
    rise = [i + 1 for i in range(half)]
    weights = rise + rise[::-1][ (0 if n % 2 == 0 else 1): ]
    s = sum(weights)
    return [w / s for w in weights]


def build_weo(ws, hist_years, proj_years):
    """WEO_Data: A=Year, B=Real GDP level, C=growth %, D=GDP deflator index,
    E=deflator YoY change %. Historical rows are left blank for you to paste the
    IMF WEO actuals; projected rows are chain formulas."""
    ws["A1"] = "Year"
    ws["B1"] = "Real GDP (constant prices)"
    ws["C1"] = "Real GDP growth (%)"
    ws["D1"] = "GDP Deflator Index"
    ws["E1"] = "GDP deflator YoY change (%)"
    years = hist_years + proj_years
    first_row = 2
    last_hist_row = first_row + len(hist_years) - 1
    for i, y in enumerate(years):
        r = first_row + i
        ws.cell(r, 1, y)
        if y in hist_years:
            # actuals: leave B/C/D blank (paste IMF WEO data here)
            if r > first_row:
                ws.cell(r, 5, f"=(D{r}/D{r-1}-1)*100")  # deflator YoY change
        else:
            # projected years -> chain formulas (never hardcode)
            ws.cell(r, 2, f"=B{r-1}*(1+C{r}/100)")               # Real GDP level
            ws.cell(r, 3, f"=$C${last_hist_row}")                 # hold last growth
            ws.cell(r, 4, f"=(1+AVERAGE($E${last_hist_row-3}:$E${last_hist_row})/100)*D{r-1}")


def build_sut(ws, n_products, supply_sheet, use_sheet):
    """SUT Calc: link SUPPLY/USE cells, compute per-product import share and the
    aggregate estimated import content share in C46."""
    ws["C3"] = "Supply: Import"
    ws["D3"] = "Supply: Resources"
    ws["E3"] = "Use: Construction"
    ws["F3"] = "Product import share"
    ws["G3"] = "Construction usage import amount"
    start = 4
    end = start + n_products - 1
    for r in range(start, end + 1):
        ws.cell(r, 2, f"Product {r - start + 1}")
        ws.cell(r, 3, f"='{supply_sheet}'!AR{r}")
        ws.cell(r, 4, f"='{supply_sheet}'!AS{r}")
        ws.cell(r, 5, f"='{use_sheet}'!T{r}")
        ws.cell(r, 6, f"=C{r}/D{r}")
        ws.cell(r, 7, f"=F{r}*E{r}")
    total = end + 1
    ws.cell(total, 2, "Total")
    ws.cell(total, 3, f"='{supply_sheet}'!AR{total}")
    ws.cell(total, 4, f"='{supply_sheet}'!AS{total}")
    ws.cell(total, 5, f"='{use_sheet}'!T{total}")
    ws["B46"] = "estimated import content share"
    ws["C46"] = f"=SUM(G{start}:G{end-1})/E{total}"


def _na_headers(ws, hdr_row):
    ws.cell(hdr_row, 2, "Year")
    ws.cell(hdr_row, 3, "Real GDP")
    ws.cell(hdr_row, 4, "Project Allocation (%)")
    ws.cell(hdr_row, 5, "Project Investment")
    ws.cell(hdr_row, 6, "Imports")
    ws.cell(hdr_row, 7, "GDP impact")
    ws.cell(hdr_row, 10, "Baseline GDP Growth")
    ws.cell(hdr_row + 1, 7, "I-M")
    ws.cell(hdr_row + 1, 8, "Multiplier")
    ws.cell(hdr_row + 1, 9, "% of GDP")


def _na_year_table(ws, first_row, n_years, weo_first_row, bell, proj_offset,
                   inv_cell, imp_cell, mult_cell, base=True, src_first=None):
    """Write one National-Accounts year table.
    proj_offset = index (0-based) of the first project year within the table.
    inv_cell/imp_cell/mult_cell = absolute refs to this block's assumption cells."""
    for i in range(n_years):
        r = first_row + i
        # Year + Real GDP link
        if base:
            ws.cell(r, 3, f"=WEO_Data!B{weo_first_row + i}*1000")  # real GDP level
        else:
            ws.cell(r, 2, f"=B{src_first + i}")
            ws.cell(r, 3, f"=C{src_first + i}")
        # baseline growth (col J) from second row on
        if i > 0:
            ws.cell(r, 10, f"=100*((C{r}/C{r-1})-1)")
    # project (shock) years: bell allocation + impact formulas
    weo_base_row = weo_first_row + proj_offset  # deflator base = shock start year
    for k, w in enumerate(bell):
        r = first_row + proj_offset + k
        weo_r = weo_base_row + k
        ws.cell(r, 4, w)                                              # allocation
        ws.cell(r, 5, f"=D{r}*{inv_cell}*(WEO_Data!$D${weo_base_row}/WEO_Data!D{weo_r})")
        ws.cell(r, 6, f"=E{r}*{imp_cell}")                            # imports
        ws.cell(r, 7, f"=E{r}-F{r}")                                  # I-M
        ws.cell(r, 8, f"=G{r}*{mult_cell}")                           # multiplier
        ws.cell(r, 9, f"=H{r}/C{r}*100")                              # % of GDP


def build_na(ws, args, bell):
    """Fill/overlay the NA sheet: baseline table + 2 scenario replicas + their
    assumption blocks. Preserves any pre-existing year rows/labels."""
    n_years = args.na_years
    proj_offset = args.proj_start_year - args.na_start_year  # e.g. 2026-2020 = 6
    weo_first = 2  # WEO_Data year rows start at row 2

    # ---- Baseline (Scenario 1) ----
    _na_headers(ws, 1)
    # ensure Year column present (template usually already has it)
    for i in range(n_years):
        r = 3 + i
        if ws.cell(r, 2).value is None:
            ws.cell(r, 2, args.na_start_year + i)
    _na_year_table(ws, 3, n_years, weo_first, bell, proj_offset,
                   "$D$30", "$D$31", "$D$32", base=True)
    ws["C28"] = "Assumptions"
    ws["C30"] = "total investment"; ws["D30"] = f"={args.investment}*{args.fx}"
    ws["C31"] = "import content share"; ws["D31"] = "='SUT Calc'!C46"
    ws["C32"] = "demand multiplier"; ws["D32"] = args.multiplier
    ws["C33"] = "Project allocation"; ws["D33"] = "Bell Shape"
    ws["E30"] = ">investment (USD) * Lari/USD FX rate"

    # ---- Scenario 2 (higher multiplier) ----
    s2 = 40
    _na_headers(ws, s2)
    _na_year_table(ws, s2 + 2, n_years, weo_first, bell, proj_offset,
                   "$D$69", "$D$70", "$D$71", base=False, src_first=3)
    ws["D67"] = "Scenario 2 (Higher Multiplier)"
    ws["C69"] = "total investment"; ws["D69"] = "=D30"
    ws["C70"] = "import content share"; ws["D70"] = "=D31"
    ws["C71"] = "demand multiplier"; ws["D71"] = args.scenario2_multiplier
    ws["C72"] = "Project allocation"; ws["D72"] = "=D33"

    # ---- Scenario 3 (higher import content share) ----
    s3 = 77
    _na_headers(ws, s3)
    _na_year_table(ws, s3 + 2, n_years, weo_first, bell, proj_offset,
                   "$D$106", "$D$107", "$D$108", base=False, src_first=3)
    ws["D104"] = "Scenario 3 (Higher Import Content Share)"
    ws["C106"] = "total investment"; ws["D106"] = "=D30"
    ws["C107"] = "import content share"; ws["D107"] = args.scenario3_import_share
    ws["C108"] = "demand multiplier"; ws["D108"] = args.multiplier
    ws["C109"] = "Project allocation"; ws["D109"] = "=D33"


def main():
    p = argparse.ArgumentParser(description="Scaffold a demand-side macro shock-analysis workbook.")
    p.add_argument("--out", default="test_demand.xlsx")
    p.add_argument("--investment", type=float, default=6500, help="shock size in source currency (e.g. USD millions/billions per task)")
    p.add_argument("--fx", type=float, default=2.746, help="local/USD exchange rate")
    p.add_argument("--multiplier", type=float, default=0.8, help="baseline demand multiplier")
    p.add_argument("--scenario2-multiplier", type=float, default=1.0)
    p.add_argument("--scenario3-import-share", type=float, default=0.5)
    p.add_argument("--na-start-year", type=int, default=2020)
    p.add_argument("--na-years", type=int, default=23, help="rows of years in the NA table (2020-2042)")
    p.add_argument("--proj-start-year", type=int, default=2026)
    p.add_argument("--proj-years", type=int, default=8, help="number of shock/project years")
    p.add_argument("--weo-hist-end", type=int, default=2027)
    p.add_argument("--weo-proj-end", type=int, default=2033)
    p.add_argument("--sut-products", type=int, default=38)
    args = p.parse_args()

    supply_sheet = "SUPPLY (38-38)-2024"
    use_sheet = "USE (38-38)-2024"

    import os
    if os.path.exists(args.out):
        wb = load_workbook(args.out)
    else:
        wb = Workbook()
        wb.active.title = "NA"

    def sheet(name):
        for s in wb.sheetnames:
            if s.lower().strip() == name.lower().strip():
                return wb[s]
        return wb.create_sheet(name)

    na = sheet("NA")
    weo = sheet("WEO_Data")
    sut = sheet("SUT Calc")
    sheet(supply_sheet)  # blank - paste Geostat SUPPLY (38-38) here
    sheet(use_sheet)     # blank - paste Geostat USE (38-38) here

    hist_years = list(range(args.na_start_year, args.weo_hist_end + 1))
    proj_years = list(range(args.weo_hist_end + 1, args.weo_proj_end + 1))
    build_weo(weo, hist_years, proj_years)
    build_sut(sut, args.sut_products, supply_sheet, use_sheet)
    bell = _bell(args.proj_years)
    build_na(na, args, bell)

    wb.save(args.out)
    print(f"Scaffolded {args.out} with sheets: {wb.sheetnames}")
    print("Next: paste collected IMF-WEO actuals (WEO_Data B/C/D) and the Geostat "
          "SUPPLY/USE tables into the blank sheets, then run: python recalc.py " + args.out)


if __name__ == "__main__":
    main()
