"""The spreadsheet preview must not build PyArrow-backed string columns.

Run 30634898569 died with SIGSEGV inside pyarrow, reached from the rollout thread pool:

    Current thread (most recent call first):
      File ".../pandas/core/arrays/string_arrow.py", line 241 in _from_sequence
      ...
      File ".../pandas/io/excel/_base.py", line 1780 in parse
      File ".../adapter.py", line ... in _spreadsheet_preview
      File ".../cap_evolve/trials.py", line 49 in _one          <- ThreadPoolExecutor worker

pandas 3.x makes `str` columns ArrowStringArray by default, so `read_excel` goes through
pyarrow's C++ layer. A segfault is uncatchable, so one bad preview kills the whole algorithm
process and every completed iteration with it (68 minutes and ~$6 in that run).

These tests pin the invariant that keeps the crashing frame out of the call stack, and that
the preview text is unchanged so benchmark results stay comparable.
"""

import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parents[2]
ADAPTER_DIR = REPO / "templates" / "adapters" / "spreadsheetbench"

pd = pytest.importorskip("pandas", reason="pandas not installed in this environment")
pytest.importorskip("openpyxl", reason="openpyxl not installed in this environment")


def _load_adapter_module():
    import importlib.util
    for p in (REPO / "core", ADAPTER_DIR, ADAPTER_DIR.parent):
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))
    spec = importlib.util.spec_from_file_location("_sb_adapter_pandas", ADAPTER_DIR / "adapter.py")
    assert spec and spec.loader
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _workbook(tmp_path: Path) -> Path:
    """A workbook with the shapes that exercise string construction: text, numbers, gaps."""
    import openpyxl
    wb = openpyxl.Workbook()
    for s in range(3):
        ws = wb.create_sheet(f"S{s}") if s else wb.active
        for r in range(1, 40):
            ws.cell(r, 1, f"lbl-{r}")
            ws.cell(r, 2, r * 1.5 if r % 3 else None)
            ws.cell(r, 3, None if r % 5 == 0 else f"text {r}")
            ws.cell(r, 4, r)
    p = tmp_path / "m.xlsx"
    wb.save(p)
    return p


def test_preview_does_not_produce_arrow_string_columns(tmp_path):
    """The crashing frame is ArrowStringArray._from_sequence — it must not be reachable."""
    mod = _load_adapter_module()
    book = _workbook(tmp_path)
    mod._spreadsheet_preview(book, 5)          # configures pandas as a side effect

    frame = mod._pandas().read_excel(book)
    backends = {type(frame[c].array).__name__ for c in frame.columns}
    assert "ArrowStringArray" not in backends, (
        f"pyarrow-backed strings are back — the SIGSEGV path is reachable again: {backends}"
    )


def test_pandas_string_storage_is_python(tmp_path):
    mod = _load_adapter_module()
    mod._spreadsheet_preview(_workbook(tmp_path), 5)
    assert mod._pandas().options.mode.string_storage == "python"


def test_preview_text_is_unchanged_by_the_backend_switch(tmp_path):
    """Results must stay comparable: the agent's prompt may not change."""
    mod = _load_adapter_module()
    book = _workbook(tmp_path)
    ours = mod._spreadsheet_preview(book, 5)

    # Recompute the same preview under pandas' DEFAULT backend for comparison.
    with pd.option_context("mode.string_storage", "pyarrow"):
        xf = pd.ExcelFile(book)
        chunks = []
        for name in xf.sheet_names:
            df = xf.parse(name)
            n = 5 if df.shape[0] > 5 else df.shape[0]
            chunks.append(f"Sheet Name: {name}\n{df.head(n).to_string()}\n" + "-" * 50)
        default = "\n".join(chunks)

    assert ours == default, "preview text changed — the agent would see a different prompt"


def test_preview_is_configured_once_not_per_call(tmp_path):
    """Per-call pd.option_context would race in the rollout thread pool; assert we don't.

    Checked by AST so prose *about* option_context (the helper's docstring explains why it is
    avoided) does not trip the assertion — only a real call does.
    """
    import ast
    tree = ast.parse((ADAPTER_DIR / "adapter.py").read_text(encoding="utf-8"))
    calls = [
        n for n in ast.walk(tree)
        if isinstance(n, ast.Call)
        and ((isinstance(n.func, ast.Attribute) and n.func.attr == "option_context")
             or (isinstance(n.func, ast.Name) and n.func.id == "option_context"))
    ]
    assert not calls, (
        "option_context mutates a process-global option and restores it on exit; in a thread "
        f"pool one thread can restore the Arrow backend while another is mid-read "
        f"(found call at line {calls[0].lineno if calls else '?'})"
    )
    mod = _load_adapter_module()
    mod._spreadsheet_preview(_workbook(tmp_path), 5)
    assert mod._pandas_configured is True


def test_concurrent_previews_all_succeed(tmp_path):
    """The real shape: the same workbook previewed from several threads at once."""
    from concurrent.futures import ThreadPoolExecutor
    mod = _load_adapter_module()
    book = _workbook(tmp_path)
    with ThreadPoolExecutor(max_workers=4) as ex:
        out = list(ex.map(lambda _: mod._spreadsheet_preview(book, 5), range(24)))
    assert len(out) == 24
    assert all(o == out[0] for o in out), "previews of one workbook must be deterministic"
    assert "Sheet Name: S1" in out[0]
